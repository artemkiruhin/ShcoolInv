from fastapi import HTTPException
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from backend.core import *

def apply_header_style(ws, row, headers, header_font, header_fill, header_alignment, thin_border):
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = 15


def apply_cell_style(cell, thin_border, wrap_text=True):
    cell.border = thin_border
    if wrap_text:
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def generate_users_report(ws, db, title_font, header_font, header_fill, header_alignment, thin_border):
    ws.title = "Users Report"

    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value="Users Report")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center")

    timestamp_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    headers = ["ID", "Username", "Email", "Full Name", "Phone Number", "Admin", "Active", "Registered At"]
    apply_header_style(ws, 4, headers, header_font, header_fill, header_alignment, thin_border)

    repo = UserRepository(db)
    users = repo.get_all()

    for i, user in enumerate(users, start=5):
        cells = [
            ws.cell(row=i, column=1, value=user.id),
            ws.cell(row=i, column=2, value=user.username),
            ws.cell(row=i, column=3, value=user.email),
            ws.cell(row=i, column=4, value=user.full_name),
            ws.cell(row=i, column=5, value=user.phone_number),
            ws.cell(row=i, column=6, value="Yes" if user.is_admin else "No"),
            ws.cell(row=i, column=7, value="Yes" if user.is_active else "No"),
            ws.cell(row=i, column=8, value=user.registered_at.strftime('%Y-%m-%d %H:%M:%S')),
        ]

        for cell in cells:
            apply_cell_style(cell, thin_border)

    for col_idx in range(1, 9):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value and not isinstance(cell, MergedCell):
                    max_length = max(max_length, len(str(cell.value)))

        adjusted_width = max(max_length, 12)
        ws.column_dimensions[column_letter].width = adjusted_width


def generate_rooms_report(ws, db, title_font, header_font, header_fill, header_alignment, thin_border):
    ws.title = "Rooms Report"

    ws.merge_cells('A1:C1')
    title_cell = ws.cell(row=1, column=1, value="Rooms Report")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center")

    timestamp_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    headers = ["ID", "Name", "Description"]
    apply_header_style(ws, 4, headers, header_font, header_fill, header_alignment, thin_border)

    repo = RoomRepository(db)
    rooms = repo.get_all()

    for i, room in enumerate(rooms, start=5):
        cells = [
            ws.cell(row=i, column=1, value=room.id),
            ws.cell(row=i, column=2, value=room.name),
            ws.cell(row=i, column=3, value=room.description or ""),
        ]

        for cell in cells:
            apply_cell_style(cell, thin_border)

    for col_idx in range(1, 4):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value and not isinstance(cell, MergedCell):
                    max_length = max(max_length, len(str(cell.value)))

        adjusted_width = max(max_length, 12)
        ws.column_dimensions[column_letter].width = adjusted_width


def generate_categories_report(ws, db, title_font, header_font, header_fill, header_alignment, thin_border):
    ws.title = "Categories Report"

    ws.merge_cells('A1:D1')
    title_cell = ws.cell(row=1, column=1, value="Inventory Categories Report")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center")

    timestamp_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    headers = ["ID", "Name", "Short Name", "Description"]
    apply_header_style(ws, 4, headers, header_font, header_fill, header_alignment, thin_border)

    repo = InventoryCategoryRepository(db)
    categories = repo.get_all()

    for i, category in enumerate(categories, start=5):
        cells = [
            ws.cell(row=i, column=1, value=category.id),
            ws.cell(row=i, column=2, value=category.name),
            ws.cell(row=i, column=3, value=category.short_name),
            ws.cell(row=i, column=4, value=category.description or ""),
        ]

        for cell in cells:
            apply_cell_style(cell, thin_border)

    for col_idx in range(1, 5):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value and not isinstance(cell, MergedCell):
                    max_length = max(max_length, len(str(cell.value)))

        adjusted_width = max(max_length, 12)
        ws.column_dimensions[column_letter].width = adjusted_width


def generate_inventory_items_report(ws, db, title_font, header_font, header_fill, header_alignment, thin_border):
    ws.title = "Inventory Items Report"

    ws.merge_cells('A1:I1')
    title_cell = ws.cell(row=1, column=1, value="Inventory Items Report")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center")

    timestamp_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    headers = ["ID", "Inventory Number", "Name", "Category", "Condition", "Room", "User", "Purchase Date",
               "Purchase Price"]
    apply_header_style(ws, 4, headers, header_font, header_fill, header_alignment, thin_border)

    repo = InventoryItemRepository(db)
    items = repo.get_all()

    category_repo = InventoryCategoryRepository(db)
    room_repo = RoomRepository(db)
    user_repo = UserRepository(db)

    for i, item in enumerate(items, start=5):
        category = category_repo.get_by_id(item.category_id)
        room = room_repo.get_by_id(item.room_id) if item.room_id else None
        user = user_repo.get_by_id(item.user_id) if item.user_id else None

        purchase_date = item.purchase_date.strftime('%Y-%m-%d') if item.purchase_date else ""

        cells = [
            ws.cell(row=i, column=1, value=item.id),
            ws.cell(row=i, column=2, value=item.inventory_number),
            ws.cell(row=i, column=3, value=item.name),
            ws.cell(row=i, column=4, value=category.name if category else ""),
            ws.cell(row=i, column=5, value=item.condition.value if item.condition else ""),
            ws.cell(row=i, column=6, value=room.name if room else ""),
            ws.cell(row=i, column=7, value=user.full_name if user else ""),
            ws.cell(row=i, column=8, value=purchase_date),
            ws.cell(row=i, column=9, value=float(item.purchase_price) if item.purchase_price else ""),
        ]

        for cell in cells:
            apply_cell_style(cell, thin_border)

    for col_idx in range(1, 10):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value and not isinstance(cell, MergedCell):
                    max_length = max(max_length, len(str(cell.value)))

        adjusted_width = max(max_length, 12)
        ws.column_dimensions[column_letter].width = adjusted_width


def generate_consumables_report(ws, db, title_font, header_font, header_fill, header_alignment, thin_border):
    ws.title = "Consumables Report"

    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value="Consumables Report")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center")

    timestamp_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    headers = ["ID", "Name", "Description", "Quantity", "Min Quantity", "Unit"]
    apply_header_style(ws, 4, headers, header_font, header_fill, header_alignment, thin_border)

    repo = ConsumableRepository(db)
    consumables = repo.get_all()

    for i, consumable in enumerate(consumables, start=5):
        cells = [
            ws.cell(row=i, column=1, value=consumable.id),
            ws.cell(row=i, column=2, value=consumable.name),
            ws.cell(row=i, column=3, value=consumable.description or ""),
            ws.cell(row=i, column=4, value=consumable.quantity),
            ws.cell(row=i, column=5, value=consumable.min_quantity),
            ws.cell(row=i, column=6, value=consumable.unit),
        ]

        if consumable.quantity <= consumable.min_quantity:
            for cell in cells:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for cell in cells:
            apply_cell_style(cell, thin_border)

    for col_idx in range(1, 7):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value and not isinstance(cell, MergedCell):
                    max_length = max(max_length, len(str(cell.value)))

        adjusted_width = max(max_length, 12)
        ws.column_dimensions[column_letter].width = adjusted_width


def generate_logs_report(ws, db, title_font, header_font, header_fill, header_alignment, thin_border):
    ws.title = "Logs Report"

    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value="System Logs Report")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center")

    timestamp_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    headers = ["ID", "Description", "Type", "Created At", "User"]
    apply_header_style(ws, 4, headers, header_font, header_fill, header_alignment, thin_border)

    repo = LogRepository(db)
    logs = repo.get_all()

    user_repo = UserRepository(db)

    log_types = {
        1: "Info",
        2: "Warning",
        3: "Error",
        4: "Critical"
    }

    for i, log in enumerate(logs, start=5):
        user = user_repo.get_by_id(log.user_id) if log.user_id else None

        cells = [
            ws.cell(row=i, column=1, value=log.id),
            ws.cell(row=i, column=2, value=log.description),
            ws.cell(row=i, column=3, value=log_types.get(log.type, str(log.type))),
            ws.cell(row=i, column=4, value=log.created_at.strftime('%Y-%m-%d %H:%M:%S')),
            ws.cell(row=i, column=5, value=user.username if user else ""),
        ]

        fill_colors = {
            1: None,
            2: PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            3: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            4: PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        }

        if log.type in fill_colors and fill_colors[log.type]:
            for cell in cells:
                cell.fill = fill_colors[log.type]

        for cell in cells:
            apply_cell_style(cell, thin_border)

    for col_idx in range(1, 6):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value and not isinstance(cell, MergedCell):
                    max_length = max(max_length, len(str(cell.value)))

        adjusted_width = max(max_length, 12)
        ws.column_dimensions[column_letter].width = adjusted_width


def generate_low_stock_report(ws, db, title_font, header_font, header_fill, header_alignment, thin_border):
    ws.title = "Low Stock Report"

    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value="Low Stock Consumables Report")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center")

    timestamp_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    headers = ["ID", "Name", "Description", "Current Quantity", "Min Quantity", "Unit"]
    apply_header_style(ws, 4, headers, header_font, header_fill, header_alignment, thin_border)

    repo = ConsumableRepository(db)
    low_stock_items = repo.get_low_stock()

    for i, item in enumerate(low_stock_items, start=5):
        cells = [
            ws.cell(row=i, column=1, value=item.id),
            ws.cell(row=i, column=2, value=item.name),
            ws.cell(row=i, column=3, value=item.description or ""),
            ws.cell(row=i, column=4, value=item.quantity),
            ws.cell(row=i, column=5, value=item.min_quantity),
            ws.cell(row=i, column=6, value=item.unit),
        ]

        for cell in cells:
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            apply_cell_style(cell, thin_border)

    summary_row = len(low_stock_items) + 6
    summary_cell = ws.cell(row=summary_row, column=1, value=f"Total low stock items: {len(low_stock_items)}")
    summary_cell.font = Font(bold=True)

    for col_idx in range(1, 7):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value and not isinstance(cell, MergedCell):
                    max_length = max(max_length, len(str(cell.value)))

        adjusted_width = max(max_length, 12)
        ws.column_dimensions[column_letter].width = adjusted_width


def generate_inventory_by_condition_report(ws, db, condition, title_font, header_font, header_fill, header_alignment,
                                           thin_border):
    ws.title = f"{condition.capitalize()} Items Report"

    ws.merge_cells('A1:I1')
    title_cell = ws.cell(row=1, column=1, value=f"Inventory Items in {condition.upper()} Condition")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center")

    timestamp_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    headers = ["ID", "Inventory Number", "Name", "Category", "Room", "User", "Purchase Date", "Purchase Price",
               "Last Updated"]
    apply_header_style(ws, 4, headers, header_font, header_fill, header_alignment, thin_border)

    repo = InventoryItemRepository(db)
    try:
        items = repo.get_by_condition(condition.upper())
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid condition: {condition}")

    category_repo = InventoryCategoryRepository(db)
    room_repo = RoomRepository(db)
    user_repo = UserRepository(db)

    for i, item in enumerate(items, start=5):
        category = category_repo.get_by_id(item.category_id)
        room = room_repo.get_by_id(item.room_id) if item.room_id else None
        user = user_repo.get_by_id(item.user_id) if item.user_id else None

        purchase_date = item.purchase_date.strftime('%Y-%m-%d') if item.purchase_date else ""
        updated_at = item.updated_at.strftime('%Y-%m-%d %H:%M:%S') if item.updated_at else ""

        cells = [
            ws.cell(row=i, column=1, value=item.id),
            ws.cell(row=i, column=2, value=item.inventory_number),
            ws.cell(row=i, column=3, value=item.name),
            ws.cell(row=i, column=4, value=category.name if category else ""),
            ws.cell(row=i, column=5, value=room.name if room else ""),
            ws.cell(row=i, column=6, value=user.full_name if user else ""),
            ws.cell(row=i, column=7, value=purchase_date),
            ws.cell(row=i, column=8, value=float(item.purchase_price) if item.purchase_price else ""),
            ws.cell(row=i, column=9, value=updated_at),
        ]

        for cell in cells:
            apply_cell_style(cell, thin_border)

    summary_row = len(items) + 6
    summary_cell = ws.cell(row=summary_row, column=1,
                           value=f"Total items in {condition.upper()} condition: {len(items)}")
    summary_cell.font = Font(bold=True)
    for col_idx in range(1, 10):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value and not isinstance(cell, MergedCell):
                    max_length = max(max_length, len(str(cell.value)))

        adjusted_width = max(max_length, 12)
        ws.column_dimensions[column_letter].width = adjusted_width
import io
from fastapi import APIRouter, Depends, HTTPException, Response
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
from enum import Enum

from backend.core.entities import (
    User, Room, InventoryCategory, InventoryItem,
    Consumable, Log, InventoryCondition
)
from backend.core.repositories import (
    UserRepository, RoomRepository, InventoryCategoryRepository,
    InventoryItemRepository, ConsumableRepository, LogRepository
)
from backend.core.schemas import (
    UserCreate, UserUpdate, UserResponse,
    RoomCreate, RoomResponse,
    InventoryCategoryCreate, InventoryCategoryResponse,
    InventoryItemCreate, InventoryItemUpdate, InventoryItemResponse,
    ConsumableCreate, ConsumableUpdate, ConsumableResponse,
    LogResponse, ReportType
)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from backend.configurations.database import get_db
from backend.services.export import *

router = APIRouter()


class InventoryCondition(str, Enum):
    NORMAL = "NORMAL"
    REQUIRES_REPAIR = "REQUIRES_REPAIR"
    WRITTEN_OFF = "WRITTEN_OFF"


@router.post("/users/", response_model=UserResponse)
def create_user(user: UserCreate, db: Session = Depends(get_db)):
    repo = UserRepository(db)
    db_user = repo.get_by_username(user.username)
    if db_user:
        raise HTTPException(status_code=400, detail="Username already registered")
    db_user = repo.get_by_email(user.email)
    if db_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    new_user = User.create(
        username=user.username,
        password_hash=user.password_hash,
        email=user.email,
        full_name=user.full_name,
        phone_number=user.phone_number,
        is_admin=user.is_admin
    )
    return repo.create(new_user)


@router.get("/users/", response_model=List[UserResponse])
def read_users(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    repo = UserRepository(db)
    return repo.get_all(skip=skip, limit=limit)


@router.get("/users/{user_id}", response_model=UserResponse)
def read_user(user_id: int, db: Session = Depends(get_db)):
    repo = UserRepository(db)
    db_user = repo.get_by_id(user_id)
    if db_user is None:
        raise HTTPException(status_code=404, detail="User not found")
    return db_user


@router.put("/users/{user_id}", response_model=UserResponse)
def update_user(user_id: int, user: UserUpdate, db: Session = Depends(get_db)):
    repo = UserRepository(db)
    db_user = repo.get_by_id(user_id)
    if db_user is None:
        raise HTTPException(status_code=404, detail="User not found")

    for key, value in user.dict(exclude_unset=True).items():
        setattr(db_user, key, value)

    return repo.update(db_user)


@router.delete("/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db)):
    repo = UserRepository(db)
    db_user = repo.get_by_id(user_id)
    if db_user is None:
        raise HTTPException(status_code=404, detail="User not found")
    repo.delete(user_id)
    return {"message": "User deleted successfully"}


# Rooms endpoints
@router.post("/rooms/", response_model=RoomResponse)
def create_room(room: RoomCreate, db: Session = Depends(get_db)):
    repo = RoomRepository(db)
    db_room = repo.get_by_name(room.name)
    if db_room:
        raise HTTPException(status_code=400, detail="Room with this name already exists")
    new_room = Room.create(name=room.name, description=room.description)
    return repo.create(new_room)


@router.get("/rooms/", response_model=List[RoomResponse])
def read_rooms(db: Session = Depends(get_db)):
    repo = RoomRepository(db)
    return repo.get_all()


@router.get("/rooms/{room_id}", response_model=RoomResponse)
def read_room(room_id: int, db: Session = Depends(get_db)):
    repo = RoomRepository(db)
    db_room = repo.get_by_id(room_id)
    if db_room is None:
        raise HTTPException(status_code=404, detail="Room not found")
    return db_room


@router.put("/rooms/{room_id}", response_model=RoomResponse)
def update_room(room_id: int, room: RoomCreate, db: Session = Depends(get_db)):
    repo = RoomRepository(db)
    db_room = repo.get_by_id(room_id)
    if db_room is None:
        raise HTTPException(status_code=404, detail="Room not found")

    db_room.name = room.name
    db_room.description = room.description

    return repo.update(db_room)


@router.delete("/rooms/{room_id}")
def delete_room(room_id: int, db: Session = Depends(get_db)):
    repo = RoomRepository(db)
    db_room = repo.get_by_id(room_id)
    if db_room is None:
        raise HTTPException(status_code=404, detail="Room not found")
    repo.delete(room_id)
    return {"message": "Room deleted successfully"}


@router.post("/inventory/categories/", response_model=InventoryCategoryResponse)
def create_category(category: InventoryCategoryCreate, db: Session = Depends(get_db)):
    repo = InventoryCategoryRepository(db)
    db_category = repo.get_by_name(category.name)
    if db_category:
        raise HTTPException(status_code=400, detail="Category with this name already exists")
    new_category = InventoryCategory.create(
        name=category.name,
        short_name=category.short_name,
        description=category.description
    )
    return repo.create(new_category)


@router.get("/inventory/categories/", response_model=List[InventoryCategoryResponse])
def read_categories(db: Session = Depends(get_db)):
    repo = InventoryCategoryRepository(db)
    return repo.get_all()


@router.get("/inventory/categories/{category_id}", response_model=InventoryCategoryResponse)
def read_category(category_id: int, db: Session = Depends(get_db)):
    repo = InventoryCategoryRepository(db)
    db_category = repo.get_by_id(category_id)
    if db_category is None:
        raise HTTPException(status_code=404, detail="Category not found")
    return db_category


@router.put("/inventory/categories/{category_id}", response_model=InventoryCategoryResponse)
def update_category(category_id: int, category: InventoryCategoryCreate, db: Session = Depends(get_db)):
    repo = InventoryCategoryRepository(db)
    db_category = repo.get_by_id(category_id)
    if db_category is None:
        raise HTTPException(status_code=404, detail="Category not found")

    db_category.name = category.name
    db_category.short_name = category.short_name
    db_category.description = category.description

    return repo.update(db_category)


@router.delete("/inventory/categories/{category_id}")
def delete_category(category_id: int, db: Session = Depends(get_db)):
    repo = InventoryCategoryRepository(db)
    db_category = repo.get_by_id(category_id)
    if db_category is None:
        raise HTTPException(status_code=404, detail="Category not found")
    repo.delete(category_id)
    return {"message": "Category deleted successfully"}


@router.post("/inventory/items/", response_model=InventoryItemResponse)
def create_item(item: InventoryItemCreate, db: Session = Depends(get_db)):
    repo = InventoryItemRepository(db)

    if item.inventory_number:
        db_item = repo.get_by_number(item.inventory_number)
        if db_item:
            raise HTTPException(status_code=400, detail="Item with this inventory number already exists")

    category_repo = InventoryCategoryRepository(db)
    db_category = category_repo.get_by_id(item.category_id)
    if not db_category:
        raise HTTPException(status_code=404, detail="Category not found")

    if item.room_id:
        room_repo = RoomRepository(db)
        db_room = room_repo.get_by_id(item.room_id)
        if not db_room:
            raise HTTPException(status_code=404, detail="Room not found")

    if item.user_id:
        user_repo = UserRepository(db)
        db_user = user_repo.get_by_id(item.user_id)
        if not db_user:
            raise HTTPException(status_code=404, detail="User not found")

    try:
        condition = InventoryCondition(item.condition.upper())
    except ValueError:
        valid_conditions = [e.value for e in InventoryCondition]
        raise HTTPException(
            status_code=400,
            detail=f"Invalid condition value. Must be one of: {', '.join(valid_conditions)}"
        )

    new_item = InventoryItem.create(
        inventory_number=item.inventory_number,
        name=item.name,
        description=item.description,
        category_id=item.category_id,
        condition=condition,
        room_id=item.room_id,
        user_id=item.user_id,
        photo=item.photo,
        purchase_date=item.purchase_date,
        purchase_price=item.purchase_price,
        warranty_until=item.warranty_until
    )

    return repo.create(new_item)


@router.get("/inventory/items/", response_model=List[InventoryItemResponse])
def read_items(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    repo = InventoryItemRepository(db)
    items = repo.get_all(skip=skip, limit=limit)
    return items


@router.get("/inventory/items/condition/{condition}", response_model=List[InventoryItemResponse])
def read_items_by_condition(condition: str, db: Session = Depends(get_db)):
    try:
        condition_enum = InventoryCondition(condition.upper())
    except ValueError:
        valid_conditions = [e.value for e in InventoryCondition]
        raise HTTPException(
            status_code=400,
            detail=f"Invalid condition. Must be one of: {', '.join(valid_conditions)}"
        )

    repo = InventoryItemRepository(db)
    items = repo.get_by_condition(condition_enum)
    return items


@router.get("/inventory/items/{item_id}", response_model=InventoryItemResponse)
def read_item(item_id: int, db: Session = Depends(get_db)):
    repo = InventoryItemRepository(db)
    db_item = repo.get_by_id(item_id)
    if db_item is None:
        raise HTTPException(status_code=404, detail="Item not found")
    return db_item


@router.put("/inventory/items/{item_id}", response_model=InventoryItemResponse)
def update_item(item_id: int, item: InventoryItemUpdate, db: Session = Depends(get_db)):
    repo = InventoryItemRepository(db)
    db_item = repo.get_by_id(item_id)
    if db_item is None:
        raise HTTPException(status_code=404, detail="Item not found")

    if item.condition is not None:
        try:
            condition = InventoryCondition(item.condition.upper())
            db_item.condition = condition
        except ValueError:
            valid_conditions = [e.value for e in InventoryCondition]
            raise HTTPException(
                status_code=400,
                detail=f"Invalid condition. Must be one of: {', '.join(valid_conditions)}"
            )

    update_data = item.dict(exclude_unset=True, exclude={"condition"})
    for key, value in update_data.items():
        setattr(db_item, key, value)

    db_item.updated_at = datetime.now()
    return repo.update(db_item)


@router.delete("/inventory/items/{item_id}")
def delete_item(item_id: int, db: Session = Depends(get_db)):
    repo = InventoryItemRepository(db)
    db_item = repo.get_by_id(item_id)
    if db_item is None:
        raise HTTPException(status_code=404, detail="Item not found")
    repo.delete(item_id)
    return {"message": "Item deleted successfully"}


@router.post("/inventory/items/{item_id}/write_off", response_model=InventoryItemResponse)
def write_off_item(item_id: int, db: Session = Depends(get_db)):
    repo = InventoryItemRepository(db)
    db_item = repo.get_by_id(item_id)
    if db_item is None:
        raise HTTPException(status_code=404, detail="Item not found")

    db_item.condition = InventoryCondition.WRITTEN_OFF
    db_item.is_written_off = True
    db_item.updated_at = datetime.now()

    return repo.update(db_item)


@router.post("/consumables/", response_model=ConsumableResponse)
def create_consumable(consumable: ConsumableCreate, db: Session = Depends(get_db)):
    repo = ConsumableRepository(db)
    new_consumable = Consumable.create(
        name=consumable.name,
        description=consumable.description,
        quantity=consumable.quantity,
        min_quantity=consumable.min_quantity,
        unit=consumable.unit
    )
    return repo.create(new_consumable)


@router.get("/consumables/", response_model=List[ConsumableResponse])
def read_consumables(db: Session = Depends(get_db)):
    repo = ConsumableRepository(db)
    return repo.get_all()


@router.get("/consumables/low_stock/", response_model=List[ConsumableResponse])
def read_low_stock_consumables(db: Session = Depends(get_db)):
    repo = ConsumableRepository(db)
    return repo.get_low_stock()


@router.get("/consumables/{consumable_id}", response_model=ConsumableResponse)
def read_consumable(consumable_id: int, db: Session = Depends(get_db)):
    repo = ConsumableRepository(db)
    db_consumable = repo.get_by_id(consumable_id)
    if db_consumable is None:
        raise HTTPException(status_code=404, detail="Consumable not found")
    return db_consumable


@router.put("/consumables/{consumable_id}", response_model=ConsumableResponse)
def update_consumable(consumable_id: int, consumable: ConsumableUpdate, db: Session = Depends(get_db)):
    repo = ConsumableRepository(db)
    db_consumable = repo.get_by_id(consumable_id)
    if db_consumable is None:
        raise HTTPException(status_code=404, detail="Consumable not found")

    for key, value in consumable.dict(exclude_unset=True).items():
        setattr(db_consumable, key, value)

    return repo.update(db_consumable)


@router.delete("/consumables/{consumable_id}")
def delete_consumable(consumable_id: int, db: Session = Depends(get_db)):
    repo = ConsumableRepository(db)
    db_consumable = repo.get_by_id(consumable_id)
    if db_consumable is None:
        raise HTTPException(status_code=404, detail="Consumable not found")
    repo.delete(consumable_id)
    return {"message": "Consumable deleted successfully"}


@router.post("/consumables/{consumable_id}/increase")
def increase_consumable(consumable_id: int, amount: int, db: Session = Depends(get_db)):
    repo = ConsumableRepository(db)
    db_consumable = repo.get_by_id(consumable_id)
    if db_consumable is None:
        raise HTTPException(status_code=404, detail="Consumable not found")
    repo.increase_quantity(consumable_id, amount)
    return {"message": f"Quantity increased by {amount}"}


@router.post("/consumables/{consumable_id}/decrease")
def decrease_consumable(consumable_id: int, amount: int, db: Session = Depends(get_db)):
    repo = ConsumableRepository(db)
    db_consumable = repo.get_by_id(consumable_id)
    if db_consumable is None:
        raise HTTPException(status_code=404, detail="Consumable not found")
    repo.decrease_quantity(consumable_id, amount)
    return {"message": f"Quantity decreased by {amount}"}


@router.get("/logs/", response_model=List[LogResponse])
def read_logs(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    repo = LogRepository(db)
    return repo.get_all(skip=skip, limit=limit)


@router.get("/logs/{log_id}", response_model=LogResponse)
def read_log(log_id: int, db: Session = Depends(get_db)):
    repo = LogRepository(db)
    db_log = repo.get_by_id(log_id)
    if db_log is None:
        raise HTTPException(status_code=404, detail="Log not found")
    return db_log

@router.get("/reports/excel")
def generate_excel_report(
        report_type: ReportType,
        condition: Optional[str] = None,
        db: Session = Depends(get_db)
):
    """
    Generate Excel reports for different data types.

    Parameters:
    - report_type: Type of report to generate
    - condition: Optional condition filter for inventory items report

    Returns:
    - Excel file as a downloadable response
    """
    wb = Workbook()
    ws = wb.active

    # Apply custom styling
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    if report_type == ReportType.USERS:
        generate_users_report(ws, db, title_font, header_font, header_fill, header_alignment, thin_border)
    elif report_type == ReportType.ROOMS:
        generate_rooms_report(ws, db, title_font, header_font, header_fill, header_alignment, thin_border)
    elif report_type == ReportType.INVENTORY_CATEGORIES:
        generate_categories_report(ws, db, title_font, header_font, header_fill, header_alignment, thin_border)
    elif report_type == ReportType.INVENTORY_ITEMS:
        generate_inventory_items_report(ws, db, title_font, header_font, header_fill, header_alignment, thin_border)
    elif report_type == ReportType.CONSUMABLES:
        generate_consumables_report(ws, db, title_font, header_font, header_fill, header_alignment, thin_border)
    elif report_type == ReportType.LOGS:
        generate_logs_report(ws, db, title_font, header_font, header_fill, header_alignment, thin_border)
    elif report_type == ReportType.LOW_STOCK:
        generate_low_stock_report(ws, db, title_font, header_font, header_fill, header_alignment, thin_border)
    elif report_type == ReportType.INVENTORY_BY_CONDITION:
        if not condition:
            raise HTTPException(status_code=400,
                                detail="Condition parameter is required for inventory_by_condition report")
        generate_inventory_by_condition_report(ws, db, condition, title_font, header_font, header_fill,
                                               header_alignment, thin_border)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )
